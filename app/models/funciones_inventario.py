from sqlalchemy import and_
from config.conexion import sesion_local
import os
import json
from datetime import datetime
from tkinter import simpledialog
import tkinter as tk
from tkinter import ttk
from openpyxl.worksheet.table import Table, TableStyleInfo
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from tkinter import messagebox, filedialog
from app.models.estructura_de_tablas import Productos, Clientes, Ventas, DetallesVenta
# ------------------- Funciones Productos -------------------
class funcionesProductos:
  

    @staticmethod
    def agregarProducto(nombre, precio, cantidad, modo=None):  

        try:

            conexion = sesion_local()

            nuevo =Productos(

             nombre = nombre,
             precio = precio,
             cantidad = cantidad
            )
            conexion.add(nuevo)
            conexion.commit()


            if modo:

                return f"datos ingresados _____ ID:{nuevo.id}___ NOMBRE:{nuevo.nombre}___PRECIO:{nuevo.precio}___CANTIDAD:{nuevo.cantidad}"

            if nuevo.id:

                messagebox.showinfo("Comprobante", f"Producto agregado correctamente ID_REGISTRO:{nuevo.id}")
            
            else:

                messagebox.showinfo("comprobante", "error al ingresar el registro a la bd")
      
        except Exception as e:
            messagebox.showerror("Error", "error inesperado codigo de error" + str(e))
        finally:
            conexion.close()
    @staticmethod
    def modificarProducto(nombre, precio, cantidad, modo = None, tree = None):

        try: 

            conexion = sesion_local()

            if modo:

                seleccion = tree.selection()

                contador = 0
 
                for x in seleccion:
                  
                  datos = tree.item(x)["values"]
                  id = datos[0]

                  
                producto = conexion.query(Productos).filter(Productos.id == id).first()

                if producto:
                       
                    if nombre not in(None,"", "escribe el nombre"):
                           
                           producto.nombre = nombre
                       
                    if precio not in(None,"", "escribe el precio"):
                           
                           producto.precio = precio

                    if cantidad not in(None,"", "escriba la cantidad"):
                           producto.cantidad = cantidad
                     

                    messagebox.showinfo("estado", "registro actualizado")   

                else:

                    messagebox.showerror("estado", "dato no encontrado")

            else:

                if tree >= 0:
                    
                     producto = conexion.query(Productos).filter(Productos == tree).first()

                     if producto:
                       
                       
                       
                       if nombre not in(None,"", "escribe el nombre"):
                           
                           producto.nombre = nombre
                       
                       if precio not in(None,"", "escribe el precio"):
                           
                           producto.precio = precio

                       if cantidad not in(None,"", "escriba la cantidad"):
                           producto.cantidad = cantidad
                    
                       messagebox.showinfo("estado", "registro actualizado")
            
                     else:

                       messagebox.showerror("estado", "error al momento de modificar")         
                     
            conexion.commit()

        except Exception as error:

            messagebox.showerror("estado","codigo de error"+ str(error))
        
        finally:

            conexion.close()
    
    def eliminarProducto(tree, modo = None):
     
     
     try:
        conexion = sesion_local()
         
       
        if modo:
           
            registro =  conexion.query(Productos).filter(Productos.id == 1).first()
       
            if registro:
               
               conexion.delete(registro)
               messagebox.showinfo("estado", "registro eliminado")
            
            else:
               
               messagebox.showinfo("estado", "error al eliminar el registro")
        else:
           
            seleccion = tree.selection()
           
            if not seleccion:
              messagebox.showwarning("Estado", "Selecciona al menos un registro")
              return

            for x in seleccion:
              valores = tree.item(x)["values"]  # ← aquí estaba el error
              id_producto = valores[0]
            
            registro =conexion.query(Productos).filter(Productos.id == id_producto).first()

            conexion.delete(registro)
            

                 
            messagebox.showinfo("Estado", f"ID:{registro.id} Producto(s) eliminado(s) correctamente")
        
        conexion.commit()

     except Exception as error:
        messagebox.showerror("Estado", "Código de error: " + str(error))

     finally:
        conexion.close()

    @staticmethod
    def mostrarProductos(tree_prod=None):
        
        conexion = sesion_local()

        try:
         
         filas = conexion.query(Productos).all()
         
         if tree_prod == None:

            products =[]

            for x in filas:
                products.append(str(x.nombre))
            
            return products
         
         else:
              for i in tree_prod.get_children():
                tree_prod.delete(i)
              for fila in filas:
                 tree_prod.insert("", "end", text=fila.id, values=(fila.id, fila.nombre, fila.precio, fila.cantidad,float(fila.precio * float(fila.cantidad))))
       
        except Exception as error:

            messagebox.showerror("estado", "error inseperado codigo de error" + str(error))   

        finally:

            conexion.close()


# ------------------- Funciones Clientes -------------------
class funcionesClientes:

    @staticmethod
    def agregarCliente(nombre, telefono, direccion ):
        conexion = sesion_local()
        
        try:
            nuevo =  Clientes(
                nombre = nombre,
                telefono = telefono,
                direccion = direccion
            )

            conexion.add(nuevo)
            conexion.commit()            
            messagebox.showinfo("Comprobante", "Cliente agregado correctamente")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo insertar el cliente:\n{e}")
        finally:
            conexion.close()

    @staticmethod
    def mostrarClientes(tree_cli=None):
        conexion = sesion_local()
        
        try:
         
         filas = conexion.query(Clientes).all()
         if tree_cli:

            for i in tree_cli.get_children():
             tree_cli.delete(i)
            for fila in filas:
             tree_cli.insert("", "end", text=fila.id, values=(fila.id, fila.nombre, fila.telefono, fila.direccion))
         
         else:
             
             return filas
             
        except Exception as error:

            messagebox.showerror("estado", "error insesperado codigo de error" + str(error))
        
        finally:

            conexion.close()

    @staticmethod
    def eliminarClientes(tree , modo= None):
        conexion = sesion_local()
        
        try:

            if modo:
                 id = simpledialog.askinteger("indice", "ingresa el id")
                 registro = conexion.query(Clientes).filter(Clientes.id == id).all()
                 for cliente in registro:
                  conexion.delete(cliente)             

            else:
                 seleccion = tree.selection()
                 if not seleccion :
                  messagebox.showerror("estado", "selecciona un registro")
                  return
                 eliminados = 0
                 for x in seleccion:

                   valores = tree.item(x)["values"]
                   indice =valores[0]
                  
                 registro = conexion.query(Clientes).filter(Clientes.id == indice).first()
                 
                 if registro:
                    
                    conexion.delete(registro)
                    messagebox.showinfo("Estado", "Cliente eliminado")

                 else:

                    messagebox.showinfo("Estado", "ID no encontrado")
                
            conexion.commit()
      
        except Exception as e:
            messagebox.showerror("Error", f"Error al eliminar cliente: {e}")
        finally:
            conexion.close()
    
    def modificarClientes(nombre, telefono, direccion, modo =None, tree = None  ):

      
        conexion =sesion_local()

        

        try:
          
            if modo:

                seleccion = tree.selection()

                if not seleccion:

                     messagebox.showwarning("estado", "selecciona un registro")
                     return
                for x in seleccion:
                 
                    datos = tree.item(x)['values']
                    indice = datos[0]

                    
                registro = conexion.query(Clientes).filter(Clientes.id == indice).first()
                if registro:
                    
                    
                    if nombre not in(None, "", "escribe el nombre"):
                      registro.nombre = nombre
                 
                    if telefono not in(None, "", "escribe el telefono"):
                      registro.telefono = telefono
                 
                    if direccion not in(None, "", "escribe la direccion"):
                      registro.direccion = direccion

                    messagebox.showinfo("estado", "registro actualizado")
                              
            else:
                indice = simpledialog.askinteger("estado", "ingresa el indice")
                registro = conexion.query(Clientes).filter(Clientes.id == indice).first()

                if registro:
                 
                 if nombre not in(None, "", "escribe el nombre"):
                    registro.nombre = nombre
                 
                 if telefono not in(None, "", "escribe el telefono"):
                    registro.telefono = telefono
                 
                 if direccion not in(None, "", "escribe la direccion"):
                    registro.direccion = direccion
                      
            conexion.commit()

        except Exception as error:
                
                messagebox.showerror("estado", "error inesperado codigo de error" + str(error))
        
        finally:
                
                conexion.close()
     

# ------------------- Funciones Reportes -------------------


class ExportadorExcel:
    CONFIG_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'config', 'config.json')

    @staticmethod
    def obtenerRutaGuardada():
        # Verificar si existe la carpeta; si no, crearla
        carpeta = os.path.dirname(ExportadorExcel.CONFIG_PATH)
        os.makedirs(carpeta, exist_ok=True)

        # Si el archivo no existe, crearlo vacío con estructura base
        if not os.path.exists(ExportadorExcel.CONFIG_PATH):
            with open(ExportadorExcel.CONFIG_PATH, 'w', encoding='utf-8') as f:
                json.dump({"ruta_excel": ""}, f)
            return ""

        # Leer el contenido del JSON
        try:
            with open(ExportadorExcel.CONFIG_PATH, 'r', encoding='utf-8') as archivo:
                config = json.load(archivo)
                return config.get("ruta_excel", "")
        except (json.JSONDecodeError, FileNotFoundError):
            # Si hay error en lectura o formato, reinicia el archivo
            with open(ExportadorExcel.CONFIG_PATH, 'w', encoding='utf-8') as f:
                json.dump({"ruta_excel": ""}, f)
            return ""

    @staticmethod
    def guardarRuta(ruta):
        with open(ExportadorExcel.CONFIG_PATH, 'w') as archivo:
            json.dump({"ruta_excel": ruta}, archivo)

    @staticmethod
    def exportar(tree_prod, tree_cli, nombre_archivo="reporte.xlsx"):
     try:
        ruta_guardada = ExportadorExcel.obtenerRutaGuardada()

        if not ruta_guardada:
            ruta_guardada = filedialog.askdirectory(title="Selecciona carpeta para guardar Excel")
            if not ruta_guardada:
                messagebox.showinfo("Cancelado", "Exportación cancelada")
                return
            ExportadorExcel.guardarRuta(ruta_guardada)

        ruta_completa = os.path.join(ruta_guardada, nombre_archivo)

        wb = Workbook()

        # Hoja de productos
        ws_prod = wb.active
        ws_prod.title = "Productos"
        columnas_prod = tree_prod["columns"]
        encabezados_prod = [tree_prod.heading(col)["text"] for col in columnas_prod]
        ws_prod.append(encabezados_prod)
        for item in tree_prod.get_children():
            valores = tree_prod.item(item)["values"]
            ws_prod.append(valores)

        # Hoja de clientes
        ws_cli = wb.create_sheet(title="Clientes")
        columnas_cli = tree_cli["columns"]
        encabezados_cli = [tree_cli.heading(col)["text"] for col in columnas_cli]
        ws_cli.append(encabezados_cli)
        for item in tree_cli.get_children():
            valores = tree_cli.item(item)["values"]
            ws_cli.append(valores)

        # Ajustar ancho de columna "Dirección" (columna 4 = D)
        ws_cli.column_dimensions["D"].width = 20

        wb.save(ruta_completa)
        messagebox.showinfo("Exportación exitosa", f"Archivo guardado en:\n{ruta_completa}")

     except Exception as e:
        messagebox.showerror("Error", f"No se pudo exportar:\n{e}")


class ventas:

    def mostrar_ventas(tree):

        conexion = sesion_local()

        registro = conexion.query(DetallesVenta).all()


        for i in tree.get_children():

            tree.delete(i)

        
        for i in registro:

            tree.insert("", "end", 
                        values = (i.id, 
                                  i.relacion_ventas.fecha_venta, 
                                  i.relacion_productos.nombre,
                                  i.cantidad,
                                  i.relacion_productos.precio,
                                  i.relacion_ventas.total
                                  ))
    
    def insertar_venta(nombre, cantidad):
      conexion = sesion_local()

      try:
        # Validación de entrada
        cantidad_tratada = float(cantidad)
        if cantidad_tratada <= 0:
            messagebox.showerror("Error", "La cantidad debe ser mayor a cero")
            return

        # Buscar producto
        registro = conexion.query(Productos).filter(Productos.nombre == nombre).first()
        if not registro:
            messagebox.showerror("Error", "Producto no encontrado")
            return

        # Validar stock
        if registro.cantidad < cantidad_tratada:
            messagebox.showinfo("Estado", f"No hay suficiente stock. Existencias: {registro.cantidad}")
            return

        # Calcular total y actualizar stock
        precio = registro.precio
        total = precio * cantidad_tratada
        registro.cantidad -= cantidad_tratada

        # Crear venta
        nueva_venta = Ventas(
            fecha_venta=datetime.now(),
            total=total
        )
        conexion.add(nueva_venta)
        conexion.flush()  # Para obtener el ID antes del commit

        # Crear detalle
        nuevo_detalle = DetallesVenta(
            venta_id=nueva_venta.id,
            producto_id=registro.id,
            cantidad=cantidad_tratada,
            precio_unitario=precio
        )
        conexion.add(nuevo_detalle)

        # Guardar cambios
        conexion.commit()
        messagebox.showinfo("Éxito", f"Venta registrada correctamente. Total: ${total:.0f}")

      except ValueError:
        messagebox.showerror("Error", "La cantidad ingresada no es válida")
      except Exception as e:
        messagebox.showerror("Error inesperado", f"Ocurrió un problema: {str(e)}")
      finally:
        conexion.close()

        








