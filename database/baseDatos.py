import sqlite3
import os
import json
from tkinter import simpledialog
import tkinter as tk
from tkinter import ttk
from openpyxl.worksheet.table import Table, TableStyleInfo
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from tkinter import messagebox, filedialog

# ------------------- Clase para crear tablas -------------------
class crearTablas:
    
    @staticmethod
    def establecerConexion():
     # Subir al directorio raíz del proyecto
     raiz_proyecto = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
     carpeta_data = os.path.join(raiz_proyecto, 'db')

     # Crear la carpeta si no existe
     if not os.path.exists(carpeta_data):
        os.makedirs(carpeta_data)

     # Ruta completa al archivo de base de datos
     ruta_db = os.path.join(carpeta_data, 'inventario.db')
     return sqlite3.connect(ruta_db)

    @staticmethod
    def crearTablas():
        conexion = crearTablas.establecerConexion()
        cursor = conexion.cursor()
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS productos(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT,
                precio REAL,
                cantidad INTEGER
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS clientes(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT,
                telefono TEXT,
                direccion TEXT
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS usuarios(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT,
                password TEXT
            )
        """)
        conexion.commit()
        conexion.close()


# ------------------- Funciones Productos -------------------
class funcionesProductos:

    @staticmethod
    def agregarProducto(nombre, precio, cantidad, modo=None):
        conexion = crearTablas.establecerConexion()
        cursor = conexion.cursor()
        try:
            cursor.execute("INSERT INTO productos(nombre, precio, cantidad) VALUES (?, ?, ?)",
                           (nombre, float(precio), int(cantidad)))
            conexion.commit()
            if modo:

                return cursor.lastrowid
            
            messagebox.showinfo("Comprobante", "Producto agregado correctamente")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo insertar el registro:\n{e}")
        finally:
            conexion.close()
    @staticmethod
    def modificarProducto(nombre, precio, cantidad, modo = None, tree = None):

        try: 

            conexion = crearTablas.establecerConexion()

            cursor = conexion.cursor()

            if modo:

                seleccion = tree.selection()

                contador = 0
 
                for x in seleccion:
                  
                  datos = tree.item(x)["values"]
                  id = datos[0]

                  cursor.execute("UPDATE productos SET nombre=?, precio=?, cantidad=? WHERE id=?",(nombre, precio, cantidad, id))

                  if cursor.rowcount >= 1:
                      
                      contador +=1
                
                if contador >= 1:

                    messagebox.showinfo("estado", "registro actualizado")   

                else:

                    messagebox.showerror("estado", "error al modificar")

            else:

                if tree >= 0:
                    
                     cursor.execute("UPDATE productos SET nombre=?, precio=?, cantidad=? WHERE id=? ", [nombre, precio, cantidad, tree])

                     if cursor.rowcount <= 1:

                       messagebox.showinfo("estado", "registro actualizado")
            
                     else:

                       messagebox.showerror("estado", "error al momento de modificar")         
                     
               

            conexion.commit()

        except Exception as error:

            messagebox.showerror("estado","codigo de error"+ str(error))
        
        finally:

            conexion.close()
    
    def eliminarProducto(tree, modo = None):
     
     
     try:
        conexion = crearTablas.establecerConexion()
        cursor = conexion.cursor() 
       
        if modo:
           
           cursor.execute("DELETE FROM productos WHERE id=?",(tree,))

           if cursor.rowcount >= 1:
               
               messagebox.showinfo("estado", "registro eliminado")
            
           else:
               
               messagebox.showinfo("estado", "error al eliminar el registro")


        else:
           
            seleccion = tree.selection()
           
            if not seleccion:
              messagebox.showwarning("Estado", "Selecciona al menos un registro")
              return

            for x in seleccion:
              valores = tree.item(x)["values"]  # ← aquí estaba el error
              id_producto = valores[0]
              cursor.execute("DELETE FROM productos WHERE id = ?", (id_producto,))
        
            
            messagebox.showinfo("Estado", "Producto(s) eliminado(s) correctamente")

        conexion.commit()

     except Exception as error:
        messagebox.showerror("Estado", "Código de error: " + str(error))

     finally:
        conexion.close()

    @staticmethod
    def mostrarProductos(tree_prod=None):
        conexion = crearTablas.establecerConexion()
        cursor = conexion.cursor()

        try:
         
         cursor.execute("SELECT * FROM productos")
         filas = cursor.fetchall()

         if tree_prod == None:

            return filas
         
         else:
              for i in tree_prod.get_children():
                tree_prod.delete(i)
              for fila in filas:
                 tree_prod.insert("", "end", text=fila[0], values=(fila[0], fila[1], fila[2], fila[3]))

        except Exception as error:

            messagebox.showerror("estado", "error inseperado codigo de error" + str(error))   

        finally:

            conexion.close()

        
    

      

# ------------------- Funciones Clientes -------------------
class funcionesClientes:

    @staticmethod
    def agregarCliente(nombre, telefono, direccion ):
        conexion = crearTablas.establecerConexion()
        cursor = conexion.cursor()
        try:
            cursor.execute("INSERT INTO clientes(nombre, telefono, direccion) VALUES(?, ?, ?)",
                           (nombre, telefono, direccion))
            conexion.commit()
             
            messagebox.showinfo("Comprobante", "Cliente agregado correctamente")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo insertar el cliente:\n{e}")
        finally:
            conexion.close()

    @staticmethod
    def mostrarClientes(tree_cli=None):
        conexion = crearTablas.establecerConexion()
        cursor = conexion.cursor()
        try:
         
         cursor.execute("SELECT * FROM clientes")
         filas = cursor.fetchall()
         if tree_cli:

            for i in tree_cli.get_children():
             tree_cli.delete(i)
            for fila in filas:
             tree_cli.insert("", "end", text=fila[0], values=(fila[0], fila[1], fila[2], fila[3]))
         
         else:
             
             return filas
             
        except Exception as error:

            messagebox.showerror("estado", "error insesperado codigo de error" + str(error))
        
        finally:

            conexion.close()

    @staticmethod
    def eliminarClientes(tree , modo= None):
        conexion = crearTablas.establecerConexion()
        cursor = conexion.cursor()
        try:

            if modo:
                 cursor.execute("DELETE FROM clientes WHERE id=?", (tree,))
     
            else:
                 seleccion = tree.selection()
                 if not seleccion :
                  messagebox.showerror("estado", "selecciona un registro")
                  return
                 eliminados = 0
                 for x in seleccion:

                   valores = tree.item(x)["values"]
                   indice =valores[0]
                   cursor.execute("DELETE FROM clientes WHERE id=?", (indice,))
              
                   if cursor.rowcount >= 1:

                     eliminados += 1
            
                 if eliminados >= 1 :
    
                    messagebox.showinfo("Estado", "Cliente eliminado")
                 else:
                    messagebox.showinfo("Estado", "ID no encontrado")
                
            conexion.commit()
      
        except Exception as e:
            messagebox.showerror("Error", f"Error al eliminar cliente: {e}")
        finally:
            conexion.close()
    
    def modificarClientes(nombre, telefono, direccion, modo =None, tree = None  ):

      
        conexion =crearTablas.establecerConexion()

        cursor = conexion.cursor()

        try:
          
            if modo:

                contador = 0

                seleccion = tree.selection()

                if not seleccion:

                     messagebox.showwarning("estado", "selecciona un registro")
                     return
                for x in seleccion:
                 
                    datos = tree.item(x)['values']
                    indice = datos[0]

                    cursor.execute("UPDATE clientes SET nombre =?, telefono =?, direccion =? WHERE id =?",(nombre, telefono, direccion, indice ))

                    if cursor.rowcount >= 1 :

                     contador += 1
            
                if contador >= 1:
                 
                     messagebox.showinfo("estado", "registro actualizado")

                 
                     
            else:
                  cursor.execute("UPDATE clientes SET nombre =?, telefono =?, direccion =? WHERE id =?",(nombre, telefono, direccion, tree))

                  messagebox.showinfo("estado", "registro actualizado")
                     
            conexion.commit()


        except Exception as error:
                
                messagebox.showerror("estado", "error inesperado codigo de error" + str(error))
        
        finally:
                
                conexion.close()
     

# ------------------- Funciones Reportes -------------------


class ExportadorExcel:
    CONFIG_PATH = os.path.join(os.path.dirname(__file__), 'database', 'config.json')

    @staticmethod
    def obtenerRutaGuardada():
        if os.path.exists(ExportadorExcel.CONFIG_PATH):
            with open(ExportadorExcel.CONFIG_PATH, 'r') as archivo:
                try:
                    config = json.load(archivo)
                    return config.get("ruta_excel", "")
                except json.JSONDecodeError:
                    return ""
        return ""

    @staticmethod
    def guardarRuta(ruta):
        with open(ExportadorExcel.CONFIG_PATH, 'w') as archivo:
            json.dump({"ruta_excel": ruta}, archivo)

    @staticmethod
    def exportar(tree_prod, tree_cli, nombre_archivo="reporte.xlsx"):
     try:
        ruta_guardada = ExportadorExcel.obtenerRutaGuardada()

        if not ruta_guardada:
            ruta_guardada = filedialog.askdirectory(title="Selecciona carpeta para guardar Excel")
            if not ruta_guardada:
                messagebox.showinfo("Cancelado", "Exportación cancelada")
                return
            ExportadorExcel.guardarRuta(ruta_guardada)

        ruta_completa = os.path.join(ruta_guardada, nombre_archivo)

        wb = Workbook()

        # Hoja de productos
        ws_prod = wb.active
        ws_prod.title = "Productos"
        columnas_prod = tree_prod["columns"]
        encabezados_prod = [tree_prod.heading(col)["text"] for col in columnas_prod]
        ws_prod.append(encabezados_prod)
        for item in tree_prod.get_children():
            valores = tree_prod.item(item)["values"]
            ws_prod.append(valores)

        # Hoja de clientes
        ws_cli = wb.create_sheet(title="Clientes")
        columnas_cli = tree_cli["columns"]
        encabezados_cli = [tree_cli.heading(col)["text"] for col in columnas_cli]
        ws_cli.append(encabezados_cli)
        for item in tree_cli.get_children():
            valores = tree_cli.item(item)["values"]
            ws_cli.append(valores)

        # Ajustar ancho de columna "Dirección" (columna 4 = D)
        ws_cli.column_dimensions["D"].width = 20

        wb.save(ruta_completa)
        messagebox.showinfo("Exportación exitosa", f"Archivo guardado en:\n{ruta_completa}")

     except Exception as e:
        messagebox.showerror("Error", f"No se pudo exportar:\n{e}")

 
# ------------------- Inicio de sesión -------------------
class inicioSesion:

    def login(self, usuario, contraseña):
        conexion = crearTablas.establecerConexion()
        cursor = conexion.cursor()
        try:
            cursor.execute("SELECT * FROM usuarios WHERE nombre=? AND password=?", (usuario, contraseña))
            resultado = cursor.fetchone()
            if resultado:
                messagebox.showinfo("Estado", "Inicio de sesión correcto")
                return True
            else:
                messagebox.showinfo("Estado", "Usuario o contraseña incorrectos")
                return False
        except Exception as e:
            messagebox.showerror("Error", f"Error inesperado: {e}")
        finally:
            conexion.close()

    @staticmethod
    def crearUsuario():
        conexion = crearTablas.establecerConexion()
        cursor = conexion.cursor()
        nombre = simpledialog.askstring("Usuario", "Ingrese nombre")
        if nombre is None:
            messagebox.showinfo("Cancelado", "Operación cancelada")
            return
        contraseña = simpledialog.askstring("Usuario", "Ingrese contraseña")
        if contraseña is None:
            messagebox.showinfo("Cancelado", "Operación cancelada")
            return
        try:
            cursor.execute("INSERT INTO usuarios(nombre, password) VALUES(?,?)", (nombre, contraseña))
            conexion.commit()
            messagebox.showinfo("Estado", "Usuario registrado correctamente")
        except Exception as e:
            messagebox.showerror("Error", f"Error al crear usuario: {e}")
        finally:
            conexion.close()


